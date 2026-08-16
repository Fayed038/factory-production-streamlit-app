"""
data_processor.py — Factory Production System Pro
Data ingestion · Cleaning · Metrics · SQLite storage · Soft delete · Audit trail · Excel/PDF export
Supports: Global Leaf format (Section col, Output (numeric), Duplex & HLP Pack sheet, etc.)
"""

import os, re, logging, traceback, sqlite3, json
from datetime import date, timedelta, datetime
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

logger = logging.getLogger("factory_pro")

# ------------------------------------------------------------------ #
# CONSTANTS
# ------------------------------------------------------------------ #
DEFAULT_SHIFT_MINUTES = 720
WASTE_THRESHOLDS = {"green": 3.0, "yellow": 5.0, "orange": 8.0}

CANONICAL_COLUMNS = [
    "Date", "Shift", "Machine_Name", "Machine_Code", "Operator", "Supervisor",
    "Output_Quantity", "Wastage_Cigarette", "Wastage_Paper",
    "Wastage_Tipping_Paper", "Dust", "Stem",
    "Wastage_Shell_Blanket", "Wastage_Slide_AluFoil",
    "Wastage_AluFoil_InnerFrame", "Wastage_BOPP",
    "Stoppage_Reason", "Stoppage_Duration_Min", "Report_Type",
]

PIECE_WASTE_COLUMNS = [
    "Wastage_Paper", "Wastage_Tipping_Paper", "Dust", "Stem",
    "Wastage_Shell_Blanket", "Wastage_Slide_AluFoil",
    "Wastage_AluFoil_InnerFrame", "Wastage_BOPP",
]
WASTE_COLUMNS   = ["Wastage_Cigarette"] + PIECE_WASTE_COLUMNS
NUMERIC_COLUMNS = ["Output_Quantity"] + WASTE_COLUMNS + ["Stoppage_Duration_Min"]

# Sheets to skip entirely
SKIP_SHEET_NAMES = {
    "notes", "daily_balances", "note", "readme", "read me", "instructions",
    "opening-closing stock", "daily summary", "summary", "stock",
}

# ------------------------------------------------------------------ #
# COLUMN ALIASES — extended to cover your exact file
# ------------------------------------------------------------------ #
COLUMN_ALIASES = {
    "Date": [
        "date", "shift date", "production date", "dt",
    ],
    "Shift": [
        "shift", "shift name", "shift (a/b)",
    ],
    "Machine_Name": [
        "machine name", "machine", "section", "m/c name", "line name",
        "line", "section name",
    ],
    "Machine_Code": [
        "machine code", "machine no", "machine number", "m/c code",
        "m/c", "section code", "equipment code", "asset code",
    ],
    "Operator": [
        "operator", "operator name", "op name", "operator(s)",
        "operators (m/c : stamp/bandroll : wrapper)",
        "operators", "operators (m/c",
    ],
    "Supervisor": [
        "supervisor", "supervisor name", "sup name", "shift supervisor",
    ],
    "Output_Quantity": [
        "output", "output quantity", "production", "production qty",
        "output qty", "total output",
        "output (stick)",      # Making sheet
        "output (sticks)",
        "output(stick)",
        "output (numeric)",    # ← your clean numeric column
        "output(numeric)",
    ],
    "Wastage_Cigarette": [
        "wastage cigarette", "cigarette wastage", "wastage-cigarette",
        "cig wastage",
        "wastage cigarette (gm)",  # ← your exact header
    ],
    "Wastage_Paper": [
        "wastage paper", "paper wastage", "wastage cig. paper",
        "wastage cig paper", "wastage cigarette paper", "cig paper",
        "wastage cig. paper (gm)",
        "wastage tipping paper (gm)",  # fallback if tipping merged
    ],
    "Wastage_Tipping_Paper": [
        "wastage tipping paper", "tipping paper wastage", "tipping wastage",
        "wastage tipping paper (gm)",
    ],
    "Dust": [
        "dust", "dust waste", "dust (u)", "dust (gm)",
    ],
    "Stem": [
        "stem", "stem waste", "stem (u)", "stem (gm)",
    ],
    "Wastage_Shell_Blanket": [
        "wastage shell/blanke", "shell/blanke", "shell/blanket",
        "wastage shell / blanke (gm)",   # ← your exact header
        "wastage shell / blanke",
    ],
    "Wastage_Slide_AluFoil": [
        "wastage slide/alufoil", "slide/alufoil",
        "wastage slide (gm) [duplex]",   # ← your exact header
        "wastage slide (gm)",
    ],
    "Wastage_AluFoil_InnerFrame": [
        "wastage alufoil/innerframe", "alufoil/innerframe",
        "wastage alu-foil (gm)",         # ← your exact header
        "wastage alu-foil",
        "wastage inner frame (gm) [hlp]",
        "wastage inner frame (gm)",
    ],
    "Wastage_BOPP": [
        "wastage bopp", "bopp",
        "wastage bopp (gm)",             # ← your exact header
    ],
    "Stoppage_Reason": [
        "stoppage reason", "stop reason", "reason for stoppage",
        "downtime reason",
        "remarks / stoppage",            # ← your exact header
        "remarks/stoppage",
        "remarks / downtime", "remarks", "stoppage",
    ],
    "Stoppage_Duration_Min": [
        "stoppage duration", "stoppage time", "downtime",
        "downtime (min)", "stoppage (min)",
    ],
}

# ------------------------------------------------------------------ #
# Duration parser
# ------------------------------------------------------------------ #
_HOUR_UNIT = r"(?:hrs?|hours?|hr)"
_MIN_UNIT  = r"(?:mins?|minutes?|min)"
_HR_MIN    = re.compile(r"(\d+(?:\.\d+)?)\s*" + _HOUR_UNIT + r"\s*(?:(\d+)\s*" + _MIN_UNIT + r")?", re.I)
_HALF_HR   = re.compile(r"(?:1/2|½)\s*" + _HOUR_UNIT + r"?", re.I)
_MIN_ONLY  = re.compile(r"(?<!\d)(\d+)\s*" + _MIN_UNIT, re.I)
# e.g. "2h30m", "1h", "30m"
_COMPACT   = re.compile(r"(\d+(?:\.\d+)?)h(?:(\d+)m)?", re.I)


def extract_duration_minutes(text):
    if not text or not isinstance(text, str):
        return np.nan
    total, found = 0.0, False
    w = str(text)

    # compact: 2h30m
    for m in _COMPACT.finditer(w):
        total += float(m.group(1)) * 60 + (float(m.group(2)) if m.group(2) else 0.0)
        found = True
    w = _COMPACT.sub(" ", w)

    # "1/2 hour"
    for _ in _HALF_HR.findall(w):
        total += 30; found = True
    w = _HALF_HR.sub(" ", w)

    # "2 hours 30 min"
    for m in _HR_MIN.finditer(w):
        total += float(m.group(1)) * 60 + (float(m.group(2)) if m.group(2) else 0.0)
        found = True
    w = _HR_MIN.sub(" ", w)

    # "30 min"
    for m in _MIN_ONLY.finditer(w):
        total += float(m.group(1)); found = True

    return total if found else np.nan


# ------------------------------------------------------------------ #
# Full machine label
# ------------------------------------------------------------------ #
def full_machine_label(name, code):
    n = str(name).strip() if pd.notna(name) else ""
    c = str(code).strip() if pd.notna(code) else ""
    if n and c and n.upper() != c.upper():
        return f"{n} ({c})"
    return n or c or "Unknown"


# ------------------------------------------------------------------ #
# SQLite Production DB
# ------------------------------------------------------------------ #
class ProductionDB:
    def __init__(self, db_path: str):
        self.db_path = db_path
        self._init_db()

    def _conn(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        return conn

    def _init_db(self):
        with self._conn() as conn:
            conn.executescript("""
                CREATE TABLE IF NOT EXISTS production_data (
                    id                INTEGER PRIMARY KEY AUTOINCREMENT,
                    date              DATE,
                    machine_code      TEXT,
                    machine_name      TEXT,
                    machine_type      TEXT,
                    shift             TEXT,
                    operator          TEXT,
                    supervisor        TEXT,
                    output            INTEGER DEFAULT 0,
                    waste_cigarette   REAL DEFAULT 0,
                    waste_paper       REAL DEFAULT 0,
                    waste_tipping     REAL DEFAULT 0,
                    dust              REAL DEFAULT 0,
                    stem              REAL DEFAULT 0,
                    waste_shell       REAL DEFAULT 0,
                    waste_alufoil     REAL DEFAULT 0,
                    waste_innerframe  REAL DEFAULT 0,
                    waste_bopp        REAL DEFAULT 0,
                    total_waste       REAL DEFAULT 0,
                    waste_percentage  REAL DEFAULT 0,
                    downtime_reason   TEXT DEFAULT '',
                    downtime_minutes  REAL DEFAULT 0,
                    report_type       TEXT,
                    data_quality_flag TEXT DEFAULT 'OK',
                    entered_by        TEXT,
                    entered_at        DATETIME DEFAULT CURRENT_TIMESTAMP,
                    is_deleted        INTEGER NOT NULL DEFAULT 0,
                    deleted_by        TEXT,
                    deleted_at        DATETIME,
                    source_file       TEXT
                );
                CREATE TABLE IF NOT EXISTS data_audit_log (
                    id         INTEGER PRIMARY KEY AUTOINCREMENT,
                    username   TEXT NOT NULL,
                    action     TEXT NOT NULL,
                    record_id  INTEGER,
                    old_values TEXT,
                    new_values TEXT,
                    timestamp  DATETIME DEFAULT CURRENT_TIMESTAMP
                );
                CREATE INDEX IF NOT EXISTS idx_prod_date    ON production_data(date);
                CREATE INDEX IF NOT EXISTS idx_prod_machine ON production_data(machine_code);
                CREATE INDEX IF NOT EXISTS idx_prod_deleted ON production_data(is_deleted);
                CREATE INDEX IF NOT EXISTS idx_daudit_ts   ON data_audit_log(timestamp);
            """)

    def _audit(self, username, action, record_id=None, old_values=None, new_values=None):
        with self._conn() as conn:
            conn.execute(
                "INSERT INTO data_audit_log (username,action,record_id,old_values,new_values)"
                " VALUES (?,?,?,?,?)",
                (username, action, record_id,
                 json.dumps(old_values) if old_values else None,
                 json.dumps(new_values) if new_values else None)
            )

    def upsert_from_df(self, df: pd.DataFrame, entered_by="system", source_file="") -> int:
        def safe_int(v):
            try:
                f = float(v)
                return 0 if np.isnan(f) else int(f)
            except (TypeError, ValueError):
                return 0

        def safe_float(v):
            try:
                f = float(v)
                return 0.0 if np.isnan(f) else f
            except (TypeError, ValueError):
                return 0.0

        inserted = 0
        with self._conn() as conn:
            for _, row in df.iterrows():
                date_val = row.get("Date")
                if pd.isna(date_val):
                    date_str = None
                else:
                    try:
                        date_str = pd.Timestamp(date_val).strftime("%Y-%m-%d")
                    except Exception:
                        date_str = str(date_val)

                output_qty  = safe_int(row.get("Output_Quantity", 0))
                exists = conn.execute(
                    "SELECT id FROM production_data WHERE date=? AND machine_name=? AND shift=?"
                    " AND operator=? AND output=? AND is_deleted=0",
                    (date_str,
                     str(row.get("Machine_Name", "") or ""),
                     str(row.get("Shift", "") or ""),
                     str(row.get("Operator", "") or ""),
                     output_qty)
                ).fetchone()
                if exists:
                    continue

                total_waste = sum(safe_float(row.get(c, 0)) for c in PIECE_WASTE_COLUMNS)
                waste_pct   = (100 * total_waste / output_qty) if output_qty > 0 else 0.0
                mname       = str(row.get("Machine_Name", "") or "")
                mtype       = "Making" if any(k in mname.lower() for k in ["making", "maker", "mk"]) else "Packing"

                conn.execute("""
                    INSERT INTO production_data
                    (date,machine_code,machine_name,machine_type,shift,operator,supervisor,
                     output,waste_cigarette,waste_paper,waste_tipping,dust,stem,
                     waste_shell,waste_alufoil,waste_innerframe,waste_bopp,
                     total_waste,waste_percentage,downtime_reason,downtime_minutes,
                     report_type,data_quality_flag,entered_by,source_file)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    date_str,
                    str(row.get("Machine_Code", "") or "") or None,
                    mname or None, mtype,
                    str(row.get("Shift", "") or "") or None,
                    str(row.get("Operator", "") or "") or None,
                    str(row.get("Supervisor", "") or "") or None,
                    output_qty,
                    safe_float(row.get("Wastage_Cigarette")),
                    safe_float(row.get("Wastage_Paper")),
                    safe_float(row.get("Wastage_Tipping_Paper")),
                    safe_float(row.get("Dust")),
                    safe_float(row.get("Stem")),
                    safe_float(row.get("Wastage_Shell_Blanket")),
                    safe_float(row.get("Wastage_Slide_AluFoil")),
                    safe_float(row.get("Wastage_AluFoil_InnerFrame")),
                    safe_float(row.get("Wastage_BOPP")),
                    total_waste, waste_pct,
                    str(row.get("Stoppage_Reason", "") or "") or None,
                    safe_float(row.get("Stoppage_Duration_Min")),
                    str(row.get("Report_Type", "") or "") or None,
                    str(row.get("Data_Quality_Flag", "OK") or "OK"),
                    entered_by, source_file,
                ))
                inserted += 1
        if inserted > 0:
            self._audit(entered_by, f"INSERT_{inserted}_ROWS", None,
                        None, {"source": source_file, "count": inserted})
        return inserted

    def soft_delete_all(self, deleted_by="admin") -> int:
        with self._conn() as conn:
            count = conn.execute(
                "SELECT COUNT(*) FROM production_data WHERE is_deleted=0"
            ).fetchone()[0]
            conn.execute(
                "UPDATE production_data SET is_deleted=1, deleted_by=?, deleted_at=CURRENT_TIMESTAMP"
                " WHERE is_deleted=0", (deleted_by,)
            )
        self._audit(deleted_by, "SOFT_DELETE_ALL", None,
                    {"is_deleted": 0}, {"is_deleted": 1, "count": count})
        return count

    def get_audit_log(self, limit=500) -> list:
        with self._conn() as conn:
            rows = conn.execute(
                "SELECT * FROM data_audit_log ORDER BY timestamp DESC LIMIT ?", (limit,)
            ).fetchall()
        return [dict(r) for r in rows]

    def get_stats(self) -> dict:
        with self._conn() as conn:
            active  = conn.execute("SELECT COUNT(*) FROM production_data WHERE is_deleted=0").fetchone()[0]
            deleted = conn.execute("SELECT COUNT(*) FROM production_data WHERE is_deleted=1").fetchone()[0]
        return {"active": active, "deleted": deleted, "total": active + deleted}


# ------------------------------------------------------------------ #
# DATA PROCESSOR
# ------------------------------------------------------------------ #
class DataProcessingError(Exception):
    pass


class DataProcessor:
    def __init__(self):
        self.errors   = []
        self.warnings = []

    def read_file(self, filepath):
        try:
            sheets = pd.read_excel(filepath, sheet_name=None, header=None)
        except Exception as e:
            msg = f"Could not open '{os.path.basename(filepath)}': {e}"
            self.errors.append(msg); raise DataProcessingError(msg)

        frames = []
        for sheet_name, raw in sheets.items():
            if sheet_name.strip().lower() in SKIP_SHEET_NAMES:
                self.warnings.append(f"Sheet '{sheet_name}' skipped (reference/summary).")
                continue
            try:
                df = self._extract_table(raw, sheet_name)
                if df is not None and not df.empty:
                    df["Report_Type"] = sheet_name
                    frames.append(df)
                    self.warnings.append(
                        f"Sheet '{sheet_name}': {len(df)} rows loaded."
                    )
                else:
                    self.warnings.append(f"Sheet '{sheet_name}': no usable rows found.")
            except Exception as e:
                self.warnings.append(f"Sheet '{sheet_name}' skipped: {e}")

        if not frames:
            msg = f"No usable data in '{os.path.basename(filepath)}'."
            self.errors.append(msg); raise DataProcessingError(msg)
        return pd.concat(frames, ignore_index=True)

    def _extract_table(self, raw, sheet_name):
        """Find the header row by scoring column name matches."""
        best_idx, best_score = None, 0
        for i in range(min(15, len(raw))):
            row_vals = [str(v).strip().lower() for v in raw.iloc[i].tolist() if pd.notna(v)]
            score = 0
            for canon, aliases in COLUMN_ALIASES.items():
                all_names = [canon.lower().replace("_", " ")] + aliases
                if any(rv in all_names for rv in row_vals):
                    score += 2
                elif any(any(a in rv or rv in a for a in all_names) for rv in row_vals):
                    score += 1
            if score > best_score:
                best_score, best_idx = score, i

        if best_idx is None or best_score < 2:
            return self._extract_any_wide_format(raw, sheet_name)

        header = raw.iloc[best_idx].tolist()
        data   = raw.iloc[best_idx + 1:].reset_index(drop=True)
        data.columns = [
            str(h).strip() if pd.notna(h) else f"col_{j}"
            for j, h in enumerate(header)
        ]
        mapped = self._map_columns(data)
        if mapped["Machine_Name"].notna().sum() == 0 and mapped["Machine_Code"].notna().sum() == 0:
            wide = self._extract_any_wide_format(raw, sheet_name)
            if wide is not None and not wide.empty:
                return wide
        return mapped

    def _extract_any_wide_format(self, raw, sheet_name):
        """Try every known non-tabular layout in turn; return the first that yields rows."""
        for extractor in (self._extract_flexible_shift_blocks,
                           self._extract_wide_format,
                           self._extract_shift_numbered_format):
            result = extractor(raw, sheet_name)
            if result is not None and not result.empty:
                return result
        return None

    _DESC_MAP = [
        (re.compile(r"^output", re.I),                    "Output_Quantity"),
        (re.compile(r"^wastage\s*cigarette$", re.I),       "Wastage_Cigarette"),
        (re.compile(r"^wastage\s*cig\.?\s*paper", re.I),   "Wastage_Paper"),
        (re.compile(r"^wastage\s*cigarette\s*paper", re.I),"Wastage_Paper"),
        (re.compile(r"^wastage\s*tipping\s*paper", re.I),  "Wastage_Tipping_Paper"),
        (re.compile(r"^dust", re.I),                       "Dust"),
        (re.compile(r"^stem", re.I),                       "Stem"),
        (re.compile(r"^wastage\s*shell", re.I),            "Wastage_Shell_Blanket"),
        (re.compile(r"^wastage\s*slide", re.I),            "Wastage_Slide_AluFoil"),
        (re.compile(r"^wastage\s*alu.?foil", re.I),        "_ALUFOIL_"),
        (re.compile(r"^wastage\s*inner\s*frame", re.I),    "_INNERFRAME_"),
        (re.compile(r"^wastage\s*bopp", re.I),             "Wastage_BOPP"),
        (re.compile(r"^boxes", re.I),                      None),
        (re.compile(r"^closing", re.I),                    None),
    ]

    @classmethod
    def _map_desc_flex(cls, desc_txt):
        d = re.sub(r"\s*\([^)]*\)\s*", " ", str(desc_txt)).strip()
        for pattern, canon in cls._DESC_MAP:
            if pattern.match(d):
                return canon
        return None

    def _extract_flexible_shift_blocks(self, raw, sheet_name):
        """
        Flexible, keyword-driven parser for 'per-machine, per-shift block'
        daily logs. Handles variants where the machine/shift header is
        either:
          - "SHIFT: A" on its own row, followed by a "<Name> (<Code>)" row, or
          - "<Name> (<Code>) - SHIFT A" combined on one row.
        followed by optional "Floor Running / Supervisor / Operator" lines,
        then an "S.N | Description | Quantity | Remarks" table, then an
        optional trailing stoppage/remarks line, then a blank separator.
        Driven by keyword matching rather than fixed row offsets, so minor
        layout drift between sheets/months doesn't break it.
        """
        nrows, ncols = raw.shape
        if nrows < 5 or ncols < 3:
            return None

        date_val = pd.NaT
        for i in range(min(3, nrows)):
            for j in range(ncols):
                cell = raw.iat[i, j]
                if pd.notna(cell) and re.search(r"date", str(cell), re.I):
                    m = re.search(r"(\d{1,2}[.\-/]\d{1,2}[.\-/]\d{2,4})", str(cell))
                    if m:
                        date_val = self._parse_date(m.group(1))
                    else:
                        nxt = raw.iat[i, j + 1] if j + 1 < ncols else None
                        if pd.notna(nxt):
                            date_val = self._parse_date(str(nxt))
                    break
            if pd.notna(date_val):
                break
        if pd.isna(date_val):
            date_val = self._parse_date(sheet_name)

        combo_re  = re.compile(r"^(.*?)\s*\(([^)]+)\)\s*-\s*SHIFT\s*[:\-]?\s*([AB])\s*$", re.I)
        shift_re  = re.compile(r"^\s*SHIFT\s*[:\-]?\s*([AB])\s*$", re.I)
        name_re   = re.compile(r"^(.*?)\s*\(([^)]+)\)\s*$")
        sup_re    = re.compile(r"^supervisor[^:]*:\s*(.*)$", re.I)
        op_re     = re.compile(r"^operator[^:]*:\s*(.*)$", re.I)
        sn_hdr_re = re.compile(r"^\s*S\.?\s*N\.?\s*$", re.I)
        stop_lbl_re = re.compile(r"^(closing|remarks)\b", re.I)

        def cell_str(i, j):
            if i >= nrows or j >= ncols:
                return None
            v = raw.iat[i, j]
            return str(v).strip() if pd.notna(v) else None

        def is_int_like(v):
            if pd.isna(v):
                return False
            try:
                int(float(v))
                return True
            except (TypeError, ValueError):
                return False

        records = []
        for start_col in range(ncols):
            i = 0
            while i < nrows:
                c0 = cell_str(i, start_col)
                if not c0:
                    i += 1
                    continue

                machine_name = machine_code = shift_label = None
                m = combo_re.match(c0)
                if m:
                    machine_name, machine_code, shift_label = m.group(1).strip(), m.group(2).strip(), m.group(3).upper()
                    i += 1
                else:
                    m = shift_re.match(c0)
                    if m:
                        shift_label = m.group(1).upper()
                        # find the machine-name row within the next few rows
                        k = i + 1
                        while k < nrows and k < i + 4:
                            nm = cell_str(k, start_col)
                            if nm:
                                mm = name_re.match(nm)
                                if mm:
                                    machine_name, machine_code = mm.group(1).strip(), mm.group(2).strip()
                                else:
                                    machine_name = nm
                                i = k + 1
                                break
                            k += 1
                        else:
                            i += 1
                    else:
                        i += 1
                        continue

                if not machine_name:
                    continue

                supervisor = operator = None
                # scan metadata lines until we hit the S.N table header (max 6 rows lookahead)
                scan_limit = min(nrows, i + 6)
                while i < scan_limit:
                    v = cell_str(i, start_col)
                    if v is None:
                        i += 1
                        continue
                    if sn_hdr_re.match(v):
                        i += 1
                        break
                    sm = sup_re.match(v)
                    if sm:
                        supervisor = sm.group(1).strip() or None
                        i += 1
                        continue
                    om = op_re.match(v)
                    if om:
                        operator = om.group(1).strip() or None
                        i += 1
                        continue
                    # Floor Running / other metadata lines — ignore and continue
                    i += 1

                values = {}
                remarks_texts = []
                while i < nrows:
                    sn_val = raw.iat[i, start_col] if start_col < ncols else None
                    if not is_int_like(sn_val):
                        break
                    desc = cell_str(i, start_col + 1) if start_col + 1 < ncols else None
                    qty  = raw.iat[i, start_col + 2] if start_col + 2 < ncols else None
                    rmk  = cell_str(i, start_col + 3) if start_col + 3 < ncols else None
                    if desc:
                        canon = self._map_desc_flex(desc)
                        if canon:
                            values[canon] = qty
                    if rmk and rmk.strip().lower() not in ("stick", "sticks", "gm", "gram", "grams"):
                        remarks_texts.append(rmk.strip())
                    i += 1

                if not values:
                    continue

                # trailing stoppage / remarks line before the next blank/block
                if i < nrows:
                    trailing = cell_str(i, start_col)
                    if trailing:
                        sm = stop_lbl_re.match(trailing)
                        if sm:
                            val = cell_str(i, start_col + 1)
                            if val:
                                remarks_texts.append(val)
                            i += 1
                        elif not combo_re.match(trailing) and not shift_re.match(trailing):
                            remarks_texts.append(trailing)
                            i += 1

                while i < nrows and not cell_str(i, start_col):
                    i += 1

                rec = {
                    "Date": date_val,
                    "Shift": shift_label,
                    "Machine_Name": machine_name,
                    "Machine_Code": machine_code,
                    "Operator": operator,
                    "Supervisor": supervisor,
                    "Output_Quantity": self._to_numeric(values.get("Output_Quantity")),
                    "Wastage_Cigarette": self._to_numeric(values.get("Wastage_Cigarette")),
                    "Wastage_Paper": self._to_numeric(values.get("Wastage_Paper")),
                    "Wastage_Tipping_Paper": self._to_numeric(values.get("Wastage_Tipping_Paper")),
                    "Dust": self._to_numeric(values.get("Dust")),
                    "Stem": self._to_numeric(values.get("Stem")),
                    "Wastage_Shell_Blanket": self._to_numeric(values.get("Wastage_Shell_Blanket")),
                    "Wastage_Slide_AluFoil": self._to_numeric(values.get("Wastage_Slide_AluFoil")),
                    "Wastage_AluFoil_InnerFrame": self._to_numeric(
                        sum(v for v in (self._to_numeric(values.get("_ALUFOIL_")),
                                        self._to_numeric(values.get("_INNERFRAME_")))
                            if pd.notna(v)) or np.nan
                    ),
                    "Wastage_BOPP": self._to_numeric(values.get("Wastage_BOPP")),
                    "Stoppage_Reason": " / ".join(remarks_texts) if remarks_texts else np.nan,
                    "Stoppage_Duration_Min": np.nan,
                    "Report_Type": sheet_name,
                }
                records.append(rec)

        if not records:
            return None
        df = pd.DataFrame(records)
        for c in CANONICAL_COLUMNS:
            if c not in df.columns:
                df[c] = np.nan
        return df[CANONICAL_COLUMNS]

    def _extract_wide_format(self, raw, sheet_name):
        """
        Parse the 'per-machine block' daily report layout: each machine
        has its own vertical block, followed by a 'Field | Shift A |
        Shift B' row, then field rows (Operator, Output, Wastage..,
        Remarks), ending at a blank row. Produces one row per machine
        per shift. The date is read from the sheet name (e.g. '15.07.26').
        """
        nrows, ncols = raw.shape
        if ncols < 2:
            return None
        date_val = self._parse_date(sheet_name)

        records = []
        i = 0
        while i < nrows - 1:
            col0 = raw.iat[i, 0]
            nxt0 = raw.iat[i + 1, 0] if i + 1 < nrows else None
            if pd.notna(col0) and pd.notna(nxt0) and str(nxt0).strip().lower() == "field":
                header_txt = str(col0).strip()
                m = re.match(r"^(.*?)\s*\(([^)]+)\)\s*$", header_txt)
                if m:
                    machine_code, machine_name = m.group(1).strip(), m.group(2).strip()
                else:
                    machine_code, machine_name = header_txt, header_txt

                j = i + 2
                field_rows = []
                while j < nrows:
                    fname = raw.iat[j, 0]
                    if pd.isna(fname) or str(fname).strip() == "":
                        break
                    val_a = raw.iat[j, 1] if ncols > 1 else None
                    val_b = raw.iat[j, 2] if ncols > 2 else None
                    field_rows.append((str(fname).strip(), val_a, val_b))
                    j += 1

                fdict = {name: (a, b) for name, a, b in field_rows}

                for shift_idx, shift_label in enumerate(["A", "B"]):
                    def fval(name):
                        pair = fdict.get(name)
                        return pair[shift_idx] if pair else None

                    operator_parts = []
                    plain_op = fval("Operator")
                    if pd.notna(plain_op) and str(plain_op).strip():
                        operator_parts.append(str(plain_op).strip())
                    for role, label in [("Duplex M/C", "Duplex"), ("Bandroll M/C", "Bandroll"),
                                        ("HLP M/C", "HLP"), ("Stamp M/C", "Stamp"),
                                        ("Wrapper M/C", "Wrapper")]:
                        v = fval(role)
                        if pd.notna(v) and str(v).strip():
                            operator_parts.append(f"{label}:{str(v).strip()}")
                    operator_str = " / ".join(operator_parts) if operator_parts else np.nan

                    skip_names = {"wastage cigarette", "wastage cigarette paper",
                                  "wastage tipping paper", "wastage bopp"}
                    middle_waste = [
                        (a, b)[shift_idx] for name, a, b in field_rows
                        if name.lower().startswith("wastage") and name.lower() not in skip_names
                    ]
                    shell_blanket = middle_waste[0] if len(middle_waste) > 0 else np.nan
                    slide_alufoil = middle_waste[1] if len(middle_waste) > 1 else np.nan
                    alufoil_innerframe = middle_waste[2] if len(middle_waste) > 2 else np.nan

                    rec = {
                        "Date": date_val,
                        "Shift": shift_label,
                        "Machine_Name": machine_name,
                        "Machine_Code": machine_code,
                        "Operator": operator_str,
                        "Supervisor": fval("Supervisor"),
                        "Output_Quantity": fval("Output"),
                        "Wastage_Cigarette": fval("Wastage Cigarette"),
                        "Wastage_Paper": fval("Wastage Cigarette Paper"),
                        "Wastage_Tipping_Paper": fval("Wastage Tipping Paper"),
                        "Dust": fval("Dust"),
                        "Stem": fval("Stem"),
                        "Wastage_Shell_Blanket": shell_blanket,
                        "Wastage_Slide_AluFoil": slide_alufoil,
                        "Wastage_AluFoil_InnerFrame": alufoil_innerframe,
                        "Wastage_BOPP": fval("Wastage BOPP"),
                        "Stoppage_Reason": fval("Remarks"),
                        "Stoppage_Duration_Min": np.nan,
                        "Report_Type": sheet_name,
                    }
                    records.append(rec)

                i = j
            else:
                i += 1

        if not records:
            return None
        df = pd.DataFrame(records)
        for c in CANONICAL_COLUMNS:
            if c not in df.columns:
                df[c] = np.nan
        return df[CANONICAL_COLUMNS]

    def _extract_shift_numbered_format(self, raw, sheet_name):
        """
        Parse the 'numbered list' daily report layout: Shift A and Shift B
        each occupy their own block of columns side-by-side (marked by a
        'SHIFT: A' / 'SHIFT: B' cell), each machine has a name row, a
        'Machine: <code>' row, a 'Supervisor: X' / 'Operator: Y' row (or a
        single packed 'Operators: ...' row), then an 'S.N | Description |
        Quantity' table (Output, Wastage Cigarette, Wastage Cig. Paper,
        Wastage Tipping Paper, Dust, Stem, Closing Stock), then 'Remarks:'.
        """
        nrows, ncols = raw.shape
        if nrows < 3 or ncols < 3:
            return None

        # Date: prefer "Date: DD.MM.YY" text in the first couple of rows,
        # fall back to the sheet name.
        date_val = pd.NaT
        for i in range(min(3, nrows)):
            for j in range(ncols):
                cell = raw.iat[i, j]
                if pd.notna(cell) and re.search(r"date", str(cell), re.I):
                    m = re.search(r"(\d{1,2}[.\-/]\d{1,2}[.\-/]\d{2,4})", str(cell))
                    if m:
                        date_val = self._parse_date(m.group(1))
                        break
            if pd.notna(date_val):
                break
        if pd.isna(date_val):
            date_val = self._parse_date(sheet_name)

        # Locate each "SHIFT: A" / "SHIFT: B" marker and its column offset.
        shift_blocks = []
        for i in range(min(10, nrows)):
            for j in range(ncols):
                cell = raw.iat[i, j]
                if pd.notna(cell):
                    m = re.match(r"^\s*SHIFT\s*[:\-]?\s*([AB])\s*$", str(cell).strip(), re.I)
                    if m:
                        shift_blocks.append((j, m.group(1).upper()))
        if not shift_blocks:
            return None

        desc_map = [
            (re.compile(r"^output", re.I),                    "Output_Quantity"),
            (re.compile(r"^wastage\s*cigarette$", re.I),       "Wastage_Cigarette"),
            (re.compile(r"^wastage\s*cig\.?\s*paper", re.I),   "Wastage_Paper"),
            (re.compile(r"^wastage\s*tipping\s*paper", re.I),  "Wastage_Tipping_Paper"),
            (re.compile(r"^dust", re.I),                       "Dust"),
            (re.compile(r"^stem", re.I),                       "Stem"),
            (re.compile(r"^wastage\s*shell", re.I),            "Wastage_Shell_Blanket"),
            (re.compile(r"^wastage\s*slide", re.I),            "_SLIDE_"),
            (re.compile(r"^wastage\s*alu.?foil", re.I),        "_ALUFOIL_"),
            (re.compile(r"^wastage\s*inner\s*frame", re.I),    "_INNERFRAME_"),
            (re.compile(r"^wastage\s*bopp", re.I),             "_BOPP_TOP_"),
            (re.compile(r"^boxes", re.I),                      "_BOXES_"),
        ]

        def map_desc(desc_txt):
            d = re.sub(r"\s*\([^)]*\)\s*", " ", str(desc_txt)).strip()
            for pattern, canon in desc_map:
                if pattern.match(d):
                    return canon
            return None

        records = []
        for start_col, shift_label in shift_blocks:
            i = 0
            while i < nrows - 1:
                name_cell = raw.iat[i, start_col]
                nxt_cell  = raw.iat[i + 1, start_col] if i + 1 < nrows else None
                if (pd.notna(name_cell) and pd.notna(nxt_cell)
                        and str(nxt_cell).strip().lower().startswith("machine:")):
                    machine_name = self._normalize_machine_name_variants(str(name_cell).strip())
                    mc = re.match(r"^machine\s*:\s*(.*)$", str(nxt_cell).strip(), re.I)
                    machine_code = mc.group(1).strip() if mc else str(nxt_cell).strip()

                    j = i + 2
                    supervisor, operator = np.nan, np.nan
                    if j < nrows:
                        left = raw.iat[j, start_col]
                        right = raw.iat[j, start_col + 2] if start_col + 2 < ncols else None
                        if pd.notna(left):
                            lm = re.match(r"^(supervisor|operators?)\s*:\s*(.*)$", str(left).strip(), re.I)
                            if lm:
                                if lm.group(1).lower().startswith("operator"):
                                    operator = lm.group(2).strip()
                                else:
                                    supervisor = lm.group(2).strip()
                        if pd.notna(right):
                            rm = re.match(r"^operators?\s*:\s*(.*)$", str(right).strip(), re.I)
                            if rm:
                                operator = rm.group(1).strip()
                        j += 1

                    if j < nrows and str(raw.iat[j, start_col]).strip().upper() in ("S.N", "SN", "SL", "SL.NO"):
                        j += 1  # skip the "S.N | Description | Quantity" header row

                    values = {}
                    while j < nrows:
                        lbl = raw.iat[j, start_col]
                        if pd.isna(lbl):
                            break
                        if str(lbl).strip().lower().startswith("remarks"):
                            break
                        desc = raw.iat[j, start_col + 1] if start_col + 1 < ncols else None
                        qty  = raw.iat[j, start_col + 2] if start_col + 2 < ncols else None
                        if pd.notna(desc):
                            canon = map_desc(desc)
                            if canon:
                                values[canon] = qty
                        j += 1

                    remarks = np.nan
                    if (j < nrows and pd.notna(raw.iat[j, start_col])
                            and str(raw.iat[j, start_col]).strip().lower().startswith("remarks")):
                        remarks = raw.iat[j, start_col + 1] if start_col + 1 < ncols else np.nan
                        j += 1
                    while j < nrows and pd.isna(raw.iat[j, start_col]):
                        j += 1

                    def num(canon):
                        return self._to_numeric(values.get(canon))

                    # "Slide" -> Wastage_Slide_AluFoil; "Alu-Foil" + "Inner Frame" ->
                    # Wastage_AluFoil_InnerFrame (kept separate so nothing is double-counted).
                    slide_alufoil = num("_SLIDE_")
                    af, iframe = num("_ALUFOIL_"), num("_INNERFRAME_")
                    af_if_parts = [v for v in (af, iframe) if pd.notna(v)]
                    alufoil_innerframe = sum(af_if_parts) if af_if_parts else np.nan

                    rec = {
                        "Date": date_val,
                        "Shift": shift_label,
                        "Machine_Name": machine_name,
                        "Machine_Code": machine_code,
                        "Operator": operator,
                        "Supervisor": supervisor,
                        "Output_Quantity": num("Output_Quantity"),
                        "Wastage_Cigarette": num("Wastage_Cigarette"),
                        "Wastage_Paper": num("Wastage_Paper"),
                        "Wastage_Tipping_Paper": num("Wastage_Tipping_Paper"),
                        "Dust": num("Dust"),
                        "Stem": num("Stem"),
                        "Wastage_Shell_Blanket": num("Wastage_Shell_Blanket"),
                        "Wastage_Slide_AluFoil": slide_alufoil,
                        "Wastage_AluFoil_InnerFrame": alufoil_innerframe,
                        "Wastage_BOPP": num("_BOPP_TOP_"),
                        "Stoppage_Reason": remarks,
                        "Stoppage_Duration_Min": np.nan,
                        "Report_Type": sheet_name,
                    }
                    records.append(rec)
                    i = j
                else:
                    i += 1

        if not records:
            return None
        df = pd.DataFrame(records)
        for c in CANONICAL_COLUMNS:
            if c not in df.columns:
                df[c] = np.nan
        return df[CANONICAL_COLUMNS]

    def _map_columns(self, df):
        rename_map  = {}
        used_canon  = set()
        used_cols   = set()
        cols_norm   = {c: str(c).strip().lower() for c in df.columns}

        # Two passes: exact first, then partial
        for pass_n in range(2):
            for col, cn in cols_norm.items():
                if col in used_cols:
                    continue
                for canon, aliases in COLUMN_ALIASES.items():
                    if canon in used_canon:
                        continue
                    all_names = [canon.lower().replace("_", " ")] + aliases
                    if pass_n == 0:
                        hit = cn in all_names
                    else:
                        hit = any(a in cn or cn in a for a in all_names)
                    if hit:
                        rename_map[col] = canon
                        used_canon.add(canon)
                        used_cols.add(col)
                        break

        df = df.rename(columns=rename_map)
        # Deduplicate columns — keep first occurrence of each name
        seen = {}
        new_cols = []
        for i, c in enumerate(df.columns):
            if c not in seen:
                seen[c] = i
                new_cols.append(c)
            else:
                new_cols.append(f"_dup_{c}_{i}")
        df.columns = new_cols
        # Drop any _dup_ columns
        df = df[[c for c in df.columns if not c.startswith("_dup_")]]
        for c in CANONICAL_COLUMNS:
            if c not in df.columns:
                df[c] = np.nan
        return df[CANONICAL_COLUMNS]

    def clean_data(self, df):
        df = df.copy()
        df = df.dropna(how="all", subset=[c for c in CANONICAL_COLUMNS if c in df.columns])

        df["Date"]         = df["Date"].apply(self._parse_date)
        df["Shift"]        = df["Shift"].apply(self._normalize_shift)
        df["Machine_Name"] = df["Machine_Name"].apply(self._normalize_machine_name)
        df["Machine_Code"] = df["Machine_Code"].apply(
            lambda x: str(x).strip() if pd.notna(x) and str(x).strip() not in ("", "nan", "-") else np.nan
        )
        df["Machine_Name"] = df["Machine_Name"].astype(object)
        blank_name = df["Machine_Name"].isna()
        df.loc[blank_name, "Machine_Name"] = df.loc[blank_name, "Machine_Code"].apply(
            self._normalize_machine_name
        )
        df["Machine_Label"] = df.apply(
            lambda r: full_machine_label(r["Machine_Name"], r["Machine_Code"]), axis=1
        )

        for col in ["Operator", "Supervisor", "Stoppage_Reason"]:
            df[col] = df[col].apply(
                lambda x: str(x).strip() if pd.notna(x) else np.nan
            )
            df[col] = df[col].replace({"nan": np.nan, "": np.nan, "-": np.nan})

        for col in NUMERIC_COLUMNS:
            df[col] = df[col].apply(self._to_numeric)

        # Extract duration from Stoppage_Reason text if no duration column
        needs_dur = df["Stoppage_Duration_Min"].isna() & df["Stoppage_Reason"].notna()
        df.loc[needs_dur, "Stoppage_Duration_Min"] = (
            df.loc[needs_dur, "Stoppage_Reason"].apply(extract_duration_minutes)
        )

        # Normalize variant machine names (e.g. "Making-07 (MK-07)" -> "Making-07")
        df["Machine_Name"] = df["Machine_Name"].apply(self._normalize_machine_name_variants)

        # Drop annotation/summary rows (Note, Duplex/HLP (all), etc.)
        JUNK_NAMES = {
            "note", "note (a)", "note (b)", "notes",
            "duplex/hlp (all)", "hlp 20 (all sets)", "hlp 10 (all sets)",
            "all sections", "summary",
        }
        junk_mask = df["Machine_Name"].fillna("").str.lower().str.strip().isin(JUNK_NAMES)
        if junk_mask.sum() > 0:
            self.warnings.append(f"Dropped {junk_mask.sum()} annotation/summary row(s).")
            df = df[~junk_mask]

        # Drop rows with no machine at all
        before = len(df)
        df = df.dropna(subset=["Machine_Name"], how="all")
        if before - len(df) > 0:
            self.warnings.append(f"Dropped {before - len(df)} row(s) with no machine identifier.")

        # Drop rows with zero/NaN output AND stoppage text starts with "stopped"
        stopped_mask = (
            df["Output_Quantity"].fillna(0) == 0
        ) & (
            df["Stoppage_Reason"].fillna("").astype(str).str.lower().str.strip().str.startswith("stopped")
        )
        if stopped_mask.sum() > 0:
            self.warnings.append(
                f"Skipped {stopped_mask.sum()} row(s) marked as STOPPED with zero output."
            )
            df = df[~stopped_mask]

        df["Data_Quality_Flag"] = df.apply(self._quality_flag, axis=1)
        df = df[~df["Operator"].isin(["-", "N/A", "n/a"])].copy()
        return df.reset_index(drop=True)

    @staticmethod
    def _normalize_machine_name(val):
        if pd.isna(val): return np.nan
        s = str(val).strip()
        if s.upper() in ("", "NAN", "NONE", "-", "N/A", "STOPPED"): return np.nan
        s = re.sub(r"(\d)'S\b", r"\1", s, flags=re.I)
        s = re.sub(r"\s+", " ", s).strip()
        return s

    @staticmethod
    def _normalize_machine_name_variants(val):
        """Collapse variant spellings of the same machine to a canonical name."""
        if pd.isna(val):
            return val
        s = str(val).strip()
        sl = s.lower()
        # Making-07 variants  (MK-07, MK, was Mk-07 = renamed on floor)
        if re.search(r"making[-\s]?0?7", sl) or "mk-07" in sl or "was mk-07" in sl:
            return "Making-07"
        # Making-06 variants
        if re.search(r"making[-\s]?0?6", sl):
            return "Making-06"
        # Strip parenthetical suffixes that are just alternate names, not codes
        # e.g. "HLP 10 Set-02 (HL10-CH-2026-01-BM)" is fine; keep as-is
        # But "Making-07 (MK)" should become "Making-07" — handled above
        return s

    @staticmethod
    def _parse_date(val):
        if pd.isna(val): return pd.NaT
        if isinstance(val, (pd.Timestamp, datetime)): return pd.Timestamp(val)
        s = re.sub(r"[./]", "-", str(val).strip())
        for fmt in ("%d-%m-%y", "%d-%m-%Y", "%Y-%m-%d", "%m-%d-%Y", "%d-%b-%Y", "%d-%b-%y"):
            try: return pd.Timestamp(datetime.strptime(s, fmt))
            except ValueError: continue
        return pd.to_datetime(s, dayfirst=True, errors="coerce")

    @staticmethod
    def _normalize_shift(val):
        if pd.isna(val): return np.nan
        s = str(val).strip().upper()
        if s.startswith("A"): return "A"
        if s.startswith("B"): return "B"
        return s or np.nan

    @staticmethod
    def _to_numeric(val):
        if pd.isna(val): return np.nan
        if isinstance(val, (int, float)): return float(val)
        # Remove commas (Indian lakh format: 5,70,000)
        s = re.sub(r",", "", str(val).strip())
        s = re.sub(r"[^\d.\-]", "", s)
        try: return float(s) if s else np.nan
        except ValueError: return np.nan

    @staticmethod
    def _quality_flag(row):
        missing = [f for f in ("Date", "Shift", "Output_Quantity") if pd.isna(row.get(f))]
        return "OK" if not missing else "Missing: " + ", ".join(missing)

    def process_files(self, filepaths, existing_df=None):
        self.errors, self.warnings = [], []
        frames = []
        if existing_df is not None and not existing_df.empty:
            frames.append(existing_df)
        for fp in filepaths:
            try:
                raw = self.read_file(fp)
                frames.append(self.clean_data(raw))
            except DataProcessingError:
                continue
            except Exception as e:
                self.errors.append(f"Unexpected error on '{os.path.basename(fp)}': {e}")
                logger.error(traceback.format_exc())
        if not frames:
            raise DataProcessingError("No files could be processed successfully.")
        combined = pd.concat(frames, ignore_index=True)
        key_cols = [c for c in CANONICAL_COLUMNS
                    if c not in ("Stoppage_Reason", "Stoppage_Duration_Min")]
        combined = combined.drop_duplicates(subset=key_cols, keep="first").reset_index(drop=True)
        if "Machine_Label" not in combined.columns:
            combined["Machine_Label"] = combined.apply(
                lambda r: full_machine_label(r["Machine_Name"], r["Machine_Code"]), axis=1
            )
        return combined


# ------------------------------------------------------------------ #
# SECTION CLASSIFICATION
# ------------------------------------------------------------------ #
MAKING_KW  = ["making", "maker", "mk"]
PACKING_KW = ["pack", "packing", "duplex", "hlp", "wrapper", "seal"]

def _classify(s):
    s = str(s).strip().lower()
    if any(k in s for k in MAKING_KW):  return "Making"
    if any(k in s for k in PACKING_KW): return "Packing"
    return "Other"

def add_section_column(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["Section"] = df["Report_Type"].apply(_classify)
    unclear = df["Section"] == "Other"
    df.loc[unclear, "Section"] = df.loc[unclear, "Machine_Name"].apply(_classify)
    return df


# ------------------------------------------------------------------ #
# METRICS
# ------------------------------------------------------------------ #
def waste_pct_color(pct):
    if pd.isna(pct): return "#888888"
    if pct < WASTE_THRESHOLDS["green"]:  return "#2ECC71"
    if pct < WASTE_THRESHOLDS["yellow"]: return "#F1C40F"
    if pct < WASTE_THRESHOLDS["orange"]: return "#E67E22"
    return "#E74C3C"


def compute_metrics(df: pd.DataFrame, shift_minutes: int = DEFAULT_SHIFT_MINUTES) -> dict:
    v = df.copy()
    if "Machine_Label" not in v.columns:
        v["Machine_Label"] = v.apply(
            lambda r: full_machine_label(r.get("Machine_Name"), r.get("Machine_Code")), axis=1
        )
    v["Cig_Waste_g"] = v["Wastage_Cigarette"].fillna(0)
    v["Piece_Waste"] = v[PIECE_WASTE_COLUMNS].sum(axis=1, skipna=True)
    v["Total_Waste"] = v["Piece_Waste"]
    v["Waste_Pct"]   = np.where(
        v["Output_Quantity"] > 0,
        100 * v["Total_Waste"] / v["Output_Quantity"], np.nan
    )

    by_machine = v.groupby("Machine_Label", dropna=True).agg(
        Machine_Name    =("Machine_Name",         "first"),
        Machine_Code    =("Machine_Code",         "first"),
        Total_Output    =("Output_Quantity",       "sum"),
        Total_Waste     =("Total_Waste",           "sum"),
        Cig_Waste_g     =("Cig_Waste_g",           "sum"),
        Total_Downtime  =("Stoppage_Duration_Min", "sum"),
        Shifts_Recorded =("Date",                  "count"),
    ).reset_index()
    by_machine["Waste_Pct"] = np.where(
        by_machine["Total_Output"] > 0,
        100 * by_machine["Total_Waste"] / by_machine["Total_Output"], np.nan
    )
    planned = by_machine["Shifts_Recorded"] * shift_minutes
    avail   = np.clip(
        1 - by_machine["Total_Downtime"].fillna(0) / planned.replace(0, np.nan), 0, 1
    )
    avg_out = np.where(
        by_machine["Shifts_Recorded"] > 0,
        by_machine["Total_Output"] / by_machine["Shifts_Recorded"], 0
    )
    best    = avg_out.max() if len(avg_out) else 1
    perf    = np.clip(avg_out / best, 0, 1) if best > 0 else np.zeros(len(by_machine))
    qual    = np.clip(1 - by_machine["Waste_Pct"].fillna(0) / 100, 0, 1)
    by_machine["Availability_%"] = avail * 100
    by_machine["Performance_%"]  = perf  * 100
    by_machine["Quality_%"]      = qual  * 100
    by_machine["OEE_%"]          = avail * perf * qual * 100
    by_machine = by_machine.sort_values("OEE_%", ascending=False).reset_index(drop=True)
    by_machine["Rank"] = by_machine.index + 1

    by_shift = v.groupby("Shift", dropna=True).agg(
        Total_Output=("Output_Quantity", "sum"),
        Total_Waste =("Total_Waste",     "sum"),
    ).reset_index()
    by_shift["Waste_Pct"] = np.where(
        by_shift["Total_Output"] > 0,
        100 * by_shift["Total_Waste"] / by_shift["Total_Output"], np.nan
    )

    by_day = v.dropna(subset=["Date"]).copy()
    by_day["_date"] = by_day["Date"].dt.date
    by_day = by_day.groupby("_date").agg(
        Total_Output=("Output_Quantity", "sum"),
        Total_Waste =("Total_Waste",     "sum"),
    ).reset_index().rename(columns={"_date": "Date"})
    by_day["Waste_Pct"] = np.where(
        by_day["Total_Output"] > 0,
        100 * by_day["Total_Waste"] / by_day["Total_Output"], np.nan
    )

    waste_bd = pd.DataFrame({
        "Waste_Type": PIECE_WASTE_COLUMNS,
        "Total":      [v[c].sum(skipna=True) for c in PIECE_WASTE_COLUMNS],
    })
    waste_bd = waste_bd[waste_bd["Total"] > 0]

    stop = v.dropna(subset=["Stoppage_Reason"]).copy()
    if not stop.empty:
        downtime_detail = stop.groupby(["Machine_Label", "Stoppage_Reason"]).agg(
            Total_Downtime_Min=("Stoppage_Duration_Min", "sum"),
            Occurrences       =("Stoppage_Reason",       "count"),
        ).reset_index().sort_values("Total_Downtime_Min", ascending=False)
        downtime_detail["Total_Downtime_Hr"] = downtime_detail["Total_Downtime_Min"].fillna(0) / 60
        downtime_detail["Label"] = (
            downtime_detail["Machine_Label"] + " — " + downtime_detail["Stoppage_Reason"]
        )
    else:
        downtime_detail = pd.DataFrame(columns=[
            "Machine_Label", "Stoppage_Reason", "Total_Downtime_Min",
            "Total_Downtime_Hr", "Occurrences", "Label"
        ])

    total_output  = float(v["Output_Quantity"].sum(skipna=True))
    total_waste   = float(v["Total_Waste"].sum(skipna=True))
    total_cig_g   = float(v["Cig_Waste_g"].sum(skipna=True))
    overall_waste = (100 * total_waste / total_output) if total_output > 0 else 0.0
    days_tracked  = int(v["Date"].dt.date.nunique()) if not v["Date"].dropna().empty else 0
    avg_per_day   = total_output / days_tracked if days_tracked > 0 else 0.0
    avg_waste_day = total_waste  / days_tracked if days_tracked > 0 else 0.0
    valid_dates   = v["Date"].dropna()

    kpis = {
        "total_output":       total_output,
        "total_waste":        total_waste,
        "total_cig_waste_g":  total_cig_g,
        "overall_waste_pct":  overall_waste,
        "active_machines":    int(v["Machine_Label"].nunique()),
        "days_tracked":       days_tracked,
        "total_records":      int(len(v)),
        "flagged_records":    int((v["Data_Quality_Flag"] != "OK").sum()),
        "avg_oee":            float(by_machine["OEE_%"].mean()) if not by_machine.empty else 0.0,
        "avg_output_per_day": avg_per_day,
        "avg_waste_per_day":  avg_waste_day,
        "date_min":           valid_dates.min() if not valid_dates.empty else None,
        "date_max":           valid_dates.max() if not valid_dates.empty else None,
    }

    return {
        "by_machine": by_machine, "by_shift": by_shift, "by_day": by_day,
        "waste_breakdown": waste_bd, "downtime": downtime_detail, "kpis": kpis,
    }


# ------------------------------------------------------------------ #
# SAMPLE DATA
# ------------------------------------------------------------------ #
MACHINE_MAP = {
    "Making-01":       "M8-IND-2024-01",
    "Making-02":       "M8-EN-2017-01",
    "Making-03":       "M8-IN-2024-02",
    "Making-04":       "M8-CH-2025-01",
    "Duplex Set":      "SS-IN-2015-01-BM",
    "HLP 10's Set-01": "HL10-CH-2015-01-BM",
    "HLP 20's Set-01": "HL20-CH-2015-01-BM",
}

def generate_sample_data(n_days=7) -> pd.DataFrame:
    rng   = np.random.default_rng(42)
    stops = [
        "Cork knife: 1 hour stop for mechanical works.",
        "Hopper jam: 30 minutes stop.",
        None, None, None,
        "Power cut: 45 minutes stop.",
    ]
    rows  = []
    start = date(2026, 7, 1)
    for d in range(n_days):
        cur = start + timedelta(days=d)
        for mname, mcode in MACHINE_MAP.items():
            for shift in ["A", "B"]:
                out = max(int(rng.normal(600000, 80000)), 50000)
                rows.append({
                    "Date": cur.strftime("%d.%m.%y"), "Shift": shift,
                    "Machine_Name": mname, "Machine_Code": mcode,
                    "Operator":    rng.choice(["Rahim", "Karim", "Salam", "Jamal"]),
                    "Supervisor":  rng.choice(["Saiful Islam", "Isrofil"]),
                    "Output_Quantity":       out,
                    "Wastage_Cigarette":     round(rng.normal(20000, 4000), 0),
                    "Wastage_Paper":         round(rng.normal(1500, 300), 0),
                    "Wastage_Tipping_Paper": round(rng.normal(50, 10), 0),
                    "Dust":                  round(rng.normal(8500, 500), 0),
                    "Stem":                  round(rng.normal(5000, 800), 0),
                    "Wastage_Shell_Blanket":     0,
                    "Wastage_Slide_AluFoil":     0,
                    "Wastage_AluFoil_InnerFrame": 0,
                    "Wastage_BOPP":              0,
                    "Stoppage_Reason":       rng.choice(stops),
                    "Stoppage_Duration_Min": np.nan,
                    "Report_Type":           "Sample",
                })
    df = pd.DataFrame(rows)
    p  = DataProcessor()
    return p.clean_data(df)


# ------------------------------------------------------------------ #
# EXCEL EXPORT
# ------------------------------------------------------------------ #
def export_master_excel(df, metrics, output_path):
    wb = Workbook(); wb.remove(wb.active)
    NAVY, GOLD = "1F3864", "C9A227"
    hf    = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    gfill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
    tf    = Font(bold=True, color=NAVY, size=13)
    thin  = Side(style="thin", color="B0B0B0")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft   = Alignment(horizontal="left",   vertical="center")

    def style_header(ws, row, ncols, fill=None):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = hf; cell.fill = fill or hfill
            cell.alignment = ctr; cell.border = bdr

    def autosize(ws, ncols, maxw=50):
        for c in range(1, ncols + 1):
            L = get_column_letter(c)
            m = max((len(str(cell.value)) for cell in ws[L] if cell.value), default=0)
            ws.column_dimensions[L].width = min(max(m + 3, 10), maxw)

    def write_df(ws, d, start_row=1, flag_col=None, gold_header=False):
        for j, col in enumerate(d.columns, 1):
            ws.cell(row=start_row, column=j, value=str(col))
        style_header(ws, start_row, len(d.columns), gfill if gold_header else hfill)
        for i, (_, row) in enumerate(d.iterrows(), start_row + 1):
            flagged = flag_col and str(row.get(flag_col, "OK")) != "OK"
            for j, col in enumerate(d.columns, 1):
                v = row[col]
                if isinstance(v, pd.Timestamp):
                    v = v.strftime("%d-%m-%Y") if not pd.isna(v) else ""
                elif pd.isna(v):
                    v = ""
                cell = ws.cell(row=i, column=j, value=v)
                cell.border = bdr; cell.alignment = lft
                if flagged:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                                            fill_type="solid")
        autosize(ws, len(d.columns))
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    ws1 = wb.create_sheet("Production Data")
    export_cols = (
        ["Date", "Shift", "Machine_Label", "Machine_Name", "Machine_Code",
         "Operator", "Supervisor", "Output_Quantity", "Wastage_Cigarette"]
        + PIECE_WASTE_COLUMNS
        + ["Stoppage_Reason", "Stoppage_Duration_Min", "Report_Type", "Data_Quality_Flag"]
    )
    export_cols = [c for c in export_cols if c in df.columns]
    out_df = df[export_cols].copy().rename(columns={
        "Output_Quantity":       "Output (sticks)",
        "Wastage_Cigarette":     "Cig Waste (gm)",
        "Stoppage_Duration_Min": "Downtime (min)",
        "Machine_Label":         "Machine (Full)",
    })
    write_df(ws1, out_df, flag_col="Data_Quality_Flag")

    ws2 = wb.create_sheet("KPI Summary")
    ws2["A1"] = "PRODUCTION KPI SUMMARY"; ws2["A1"].font = tf
    kpis = metrics["kpis"]
    dr   = ""
    if kpis["date_min"] and kpis["date_max"]:
        dr = (f"{kpis['date_min'].strftime('%d.%m.%y')} to "
              f"{kpis['date_max'].strftime('%d.%m.%y')}")
    r = 3
    for label, val in [
        ("Date Range",       dr),
        ("Total Output",     f"{kpis['total_output']:,.0f} sticks"),
        ("Total Piece Waste",f"{kpis['total_waste']:,.0f} sticks"),
        ("Cigarette Waste",  f"{kpis['total_cig_waste_g']:,.0f} gm"),
        ("Overall Waste %",  f"{kpis['overall_waste_pct']:.2f}%"),
        ("Avg Output / Day", f"{kpis['avg_output_per_day']:,.0f} sticks/day"),
        ("Avg Waste / Day",  f"{kpis['avg_waste_per_day']:,.0f} sticks/day"),
        ("Average OEE",      f"{kpis['avg_oee']:.1f}%"),
        ("Active Machines",  kpis["active_machines"]),
        ("Days Tracked",     kpis["days_tracked"]),
        ("Total Records",    kpis["total_records"]),
        ("Flagged Records",  kpis["flagged_records"]),
        ("Generated",        datetime.now().strftime("%d %b %Y %H:%M")),
    ]:
        ws2.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=val); r += 1
    autosize(ws2, 2)

    ws3 = wb.create_sheet("Machine OEE")
    bm  = metrics["by_machine"][[
        "Rank", "Machine_Label", "Machine_Code", "Total_Output",
        "Total_Waste", "Waste_Pct", "Availability_%", "Performance_%", "Quality_%", "OEE_%"
    ]].copy().rename(columns={
        "Machine_Label": "Machine (Full)", "Total_Output": "Output (sticks)",
        "Total_Waste":   "Waste (sticks)", "Waste_Pct":    "Waste %",
    })
    write_df(ws3, bm)

    ws4 = wb.create_sheet("Downtime Detail")
    dt  = metrics["downtime"]
    if not dt.empty:
        dt_out = dt[["Machine_Label", "Stoppage_Reason", "Total_Downtime_Min",
                      "Total_Downtime_Hr", "Occurrences"]].copy().rename(columns={
            "Machine_Label":      "Machine (Full)",
            "Total_Downtime_Min": "Downtime (min)",
            "Total_Downtime_Hr":  "Downtime (hr)",
        })
        write_df(ws4, dt_out)
    else:
        ws4["A1"] = "No stoppage data recorded."

    ws5 = wb.create_sheet("Waste Breakdown")
    write_df(ws5, metrics["waste_breakdown"])

    ws6 = wb.create_sheet("Shift Comparison")
    bs  = metrics["by_shift"].copy().rename(columns={
        "Total_Output": "Output (sticks)", "Total_Waste": "Waste (sticks)", "Waste_Pct": "Waste %"
    })
    write_df(ws6, bs)

    ws7 = wb.create_sheet("Daily Trend")
    bd  = metrics["by_day"].copy().rename(columns={
        "Total_Output": "Output (sticks)", "Total_Waste": "Waste (sticks)", "Waste_Pct": "Waste %"
    })
    write_df(ws7, bd)

    wb.save(output_path)
    return output_path


# ------------------------------------------------------------------ #
# PDF EXPORT
# ------------------------------------------------------------------ #
def export_pdf_report(metrics, output_path):
    if not REPORTLAB_AVAILABLE:
        return None
    doc    = SimpleDocTemplate(output_path, pagesize=landscape(A4),
                               topMargin=1.5*cm, bottomMargin=1.5*cm,
                               leftMargin=1.5*cm, rightMargin=1.5*cm)
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle("T",  parent=styles["Title"],   textColor=colors.HexColor("#1F3864"))
    h2_s    = ParagraphStyle("H2", parent=styles["Heading2"], textColor=colors.HexColor("#1F3864"))
    elems   = [Paragraph("Factory Production Report", title_s), Spacer(1, 6)]

    kpis = metrics["kpis"]
    dr   = ""
    if kpis["date_min"] and kpis["date_max"]:
        dr = (f"{kpis['date_min'].strftime('%d %b %Y')} – "
              f"{kpis['date_max'].strftime('%d %b %Y')}")
    elems += [
        Paragraph(f"Date Range: {dr}", styles["Normal"]),
        Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", styles["Normal"]),
        Spacer(1, 12),
    ]

    def make_table(data, col_widths=None):
        t = Table(data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0),  colors.HexColor("#1F3864")),
            ("TEXTCOLOR",  (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",   (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 9),
            ("GRID",       (0, 0), (-1, -1), 0.4, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return t

    kpi_rows = [["Metric", "Value"]] + [
        ["Total Output",      f"{kpis['total_output']:,.0f} sticks"],
        ["Total Piece Waste", f"{kpis['total_waste']:,.0f} sticks"],
        ["Cigarette Waste",   f"{kpis['total_cig_waste_g']:,.0f} gm"],
        ["Overall Waste %",   f"{kpis['overall_waste_pct']:.2f}%"],
        ["Avg Output / Day",  f"{kpis['avg_output_per_day']:,.0f} sticks/day"],
        ["Average OEE",       f"{kpis['avg_oee']:.1f}%"],
        ["Active Machines",   str(kpis["active_machines"])],
        ["Days Tracked",      str(kpis["days_tracked"])],
    ]
    elems += [Paragraph("Key Performance Indicators", h2_s), Spacer(1, 4),
              make_table(kpi_rows, [7*cm, 7*cm]), Spacer(1, 16)]

    bm = metrics["by_machine"]
    if not bm.empty:
        oee_rows = [["Rank", "Machine", "Output (sticks)", "Waste %",
                     "Avail %", "Perf %", "Qual %", "OEE %"]]
        for _, row in bm.iterrows():
            oee_rows.append([
                int(row["Rank"]), row["Machine_Label"],
                f"{row['Total_Output']:,.0f}",
                f"{row['Waste_Pct']:.2f}%" if pd.notna(row["Waste_Pct"]) else "N/A",
                f"{row['Availability_%']:.1f}%", f"{row['Performance_%']:.1f}%",
                f"{row['Quality_%']:.1f}%",
                f"{row['OEE_%']:.1f}%"     if pd.notna(row["OEE_%"]) else "N/A",
            ])
        elems += [Paragraph("Machine OEE Ranking", h2_s), Spacer(1, 4),
                  make_table(oee_rows, [1.2*cm, 6*cm, 3*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2*cm]),
                  Spacer(1, 16)]

    dt = metrics["downtime"]
    if not dt.empty:
        dt_rows = [["Machine", "Reason", "Downtime (hr)", "Occurrences"]]
        for _, row in dt.head(15).iterrows():
            dt_rows.append([
                row["Machine_Label"], row["Stoppage_Reason"],
                f"{row['Total_Downtime_Hr']:.2f} hr", int(row["Occurrences"]),
            ])
        elems += [Paragraph("Downtime by Machine & Reason", h2_s), Spacer(1, 4),
                  make_table(dt_rows, [6*cm, 9*cm, 3*cm, 2.5*cm])]

    doc.build(elems)
    return output_path